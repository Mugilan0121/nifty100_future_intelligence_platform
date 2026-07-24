"""
Financial Screener Engine.

Sprint 3 - Day 15

Loads screener configuration and applies threshold filters
to the financial_ratios dataset.
"""

import logging
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
import yaml

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)


# ---------------------------------------------------------------------
# Project paths
# ---------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = PROJECT_ROOT / "nifty100.db"
CONFIG_PATH = PROJECT_ROOT / "config" / "screener_config.yaml"
OUTPUT_PATH = PROJECT_ROOT / "output" / "screener_output.xlsx"


# ---------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s - %(message)s",
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------
# Composite Score Weights
# ---------------------------------------------------------------------

COMPOSITE_WEIGHTS = {
    "return_on_equity_pct_score": 0.15,
    "return_on_capital_employed_pct_score": 0.10,
    "net_profit_margin_pct_score": 0.10,
    "cfo_to_pat_ratio_score": 0.10,
    "positive_fcf_score": 0.05,
    "revenue_cagr_5yr_score": 0.10,
    "pat_cagr_5yr_score": 0.10,
    "debt_to_equity_score": 0.10,
    "interest_coverage_score": 0.05,
}

EXPORT_COLUMNS = [
    "company_id",
    "broad_sector",
    "sub_sector",
    "market_cap_category",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "free_cash_flow_cr",
    "cash_from_operations_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "debt_to_equity",
    "interest_coverage",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "market_cap_crore",
    "sales",
    "composite_quality_score",
    "sector_relative_score",
]


# ---------------------------------------------------------------------
# Composite Score Helpers
# ---------------------------------------------------------------------

def winsorize_series(
    series: pd.Series,
    lower_percentile: float = 0.10,
    upper_percentile: float = 0.90,
) -> pd.Series:
    """
    Winsorize a numeric pandas Series.

    Values below the lower percentile are clipped to P10.
    Values above the upper percentile are clipped to P90.
    """

    # Ignore missing values while calculating percentiles
    valid_values = series.dropna()

    if valid_values.empty:
        return series

    lower_limit = valid_values.quantile(lower_percentile)
    upper_limit = valid_values.quantile(upper_percentile)

    return series.clip(
        lower=lower_limit,
        upper=upper_limit,
    )

def normalize_series(
    series: pd.Series,
) -> pd.Series:
    """
    Normalize a numeric Series to a 0–100 scale.

    Winsorisation should be applied before calling this function.
    """

    minimum = series.min()
    maximum = series.max()

    if pd.isna(minimum) or pd.isna(maximum):
        return pd.Series(
            0,
            index=series.index,
            dtype=float,
        )

    if maximum == minimum:
        return pd.Series(
            100,
            index=series.index,
            dtype=float,
        )

    normalized = (
        (series - minimum)
        / (maximum - minimum)
    ) * 100

    return normalized

def calculate_composite_score(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Calculate the weighted composite quality score (0–100)
    for each company.
    """

    scored_df = df.copy()

    # ----------------------------------------------------------
    # Derived Metrics
    # ----------------------------------------------------------

    # CFO / PAT Ratio
    scored_df["cfo_to_pat_ratio"] = (
        scored_df["cash_from_operations_cr"]
        / scored_df["net_profit"]
    )

    # Positive Free Cash Flow Flag
    scored_df["positive_fcf_score"] = np.where(
        scored_df["free_cash_flow_cr"] > 0,
        100,
        0,
    )

    # ----------------------------------------------------------
    # Normalize Metrics
    # ----------------------------------------------------------

    metrics = [
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "cfo_to_pat_ratio",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "interest_coverage",
    ]

    for metric in metrics:

        winsorized = winsorize_series(
            scored_df[metric]
        )

        scored_df[f"{metric}_score"] = normalize_series(
            winsorized
        )

    # ----------------------------------------------------------
    # Debt-to-Equity Score
    # Lower Debt = Higher Score
    # ----------------------------------------------------------

    debt_score = normalize_series(
        winsorize_series(
            scored_df["debt_to_equity"]
        )
    )

    scored_df["debt_to_equity_score"] = (
        100 - debt_score
    )

        # ----------------------------------------------------------
    # Composite Quality Score
    # ----------------------------------------------------------

    scored_df["composite_quality_score"] = 0.0

    for metric, weight in COMPOSITE_WEIGHTS.items():

        if metric in scored_df.columns:

            scored_df["composite_quality_score"] += (
                scored_df[metric] * weight
            )

    return scored_df

def calculate_sector_relative_score(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Calculate sector-relative composite score.

    Scores are normalized within each broad sector
    so companies are compared only against sector peers.
    """

    scored_df = df.copy()

    scored_df["sector_relative_score"] = (
        scored_df
        .groupby("broad_sector")["composite_quality_score"]
        .transform(
            lambda x: normalize_series(
                winsorize_series(x)
            )
        )
    )

    return scored_df


# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

def load_screener_config() -> dict:
    """
    Load screener configuration from YAML file.
    """

    if not CONFIG_PATH.exists():
        raise FileNotFoundError(
            f"Screener configuration not found: {CONFIG_PATH}"
        )

    with open(CONFIG_PATH, "r", encoding="utf-8") as file:
        config = yaml.safe_load(file)

    logger.info("Loaded screener configuration.")

    return config

# ---------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------

def load_financial_ratios() -> pd.DataFrame:
    """
    Load the financial_ratios table from SQLite.
    """

    if not DATABASE_PATH.exists():
        raise FileNotFoundError(
            f"Database not found: {DATABASE_PATH}"
        )

    with sqlite3.connect(DATABASE_PATH) as conn:
        df = pd.read_sql_query(
    """
    SELECT
        fr.*,

        mc.pe_ratio,
        mc.pb_ratio,
        mc.dividend_yield_pct,
        mc.market_cap_crore,

        pl.sales,
        pl.net_profit,

        s.broad_sector,
        s.sub_sector,
        s.market_cap_category

    FROM financial_ratios fr

    LEFT JOIN market_cap mc
    ON fr.company_id = mc.company_id
    AND substr(fr.year,-4)=CAST(mc.year AS TEXT)

    LEFT JOIN profitandloss pl
    ON fr.company_id = pl.company_id
    AND fr.year = pl.year

    LEFT JOIN sectors s
    ON fr.company_id = s.company_id
    """,
    conn,
)

    logger.info(
    "Loaded financial_ratios table (%d rows).",
    len(df),
)

# ----------------------------------------------------------
# Remove duplicate company-year records
# ----------------------------------------------------------

    df = df.drop_duplicates()

    logger.info(
    "After removing duplicate rows: %d records.",
    len(df),
)

# ----------------------------------------------------------
# Keep latest annual record for each company
# ----------------------------------------------------------

    df = (
    df.sort_values("year")
      .groupby("company_id", as_index=False)
      .tail(1)
      .reset_index(drop=True)
)

    logger.info(
    "Latest annual dataset contains %d companies.",
    len(df),
)

    return df

# ---------------------------------------------------------------------
# Filter Engine
# ---------------------------------------------------------------------

def apply_filters(
    df: pd.DataFrame,
    filters: dict,
) -> pd.DataFrame:
    """
    Apply threshold filters to a DataFrame.
    """

    filtered_df = df.copy()

    for column, limits in filters.items():

        if column not in filtered_df.columns:
            logger.warning(
                "Column '%s' not found. Skipping filter.",
                column,
            )
            continue

        if "min" in limits:
            filtered_df = filtered_df[
                filtered_df[column] >= limits["min"]
            ]

        if "max" in limits:
            filtered_df = filtered_df[
                filtered_df[column] <= limits["max"]
            ]

    logger.info(
        "Filtered dataframe contains %d rows.",
        len(filtered_df),
    )

    return filtered_df


def apply_de_filter(
    df: pd.DataFrame,
    max_de: float,
) -> pd.DataFrame:
    """
    Apply Debt-to-Equity filter while excluding Financial companies.
    """

    financial_df = df[
        df["broad_sector"] == "Financials"
    ]

    non_financial_df = df[
        df["broad_sector"] != "Financials"
    ]

    non_financial_df = non_financial_df[
        non_financial_df["debt_to_equity"] <= max_de
    ]

    filtered_df = pd.concat(
        [financial_df, non_financial_df],
        ignore_index=True,
    )

    logger.info(
        "Applied Debt-to-Equity filter excluding Financial sector."
    )

    return filtered_df

def apply_icr_filter(
    df: pd.DataFrame,
    min_icr: float,
) -> pd.DataFrame:
    """
    Apply Interest Coverage Ratio filter.

    Companies with missing Interest Coverage (Debt Free)
    automatically pass the filter.
    """

    filtered_df = df[
        (df["interest_coverage"].isna())
        | (df["interest_coverage"] >= min_icr)
    ]

    logger.info(
        "Applied Interest Coverage filter (Debt Free companies included)."
    )

    return filtered_df

# ---------------------------------------------------------------------
# Main Screener
# ---------------------------------------------------------------------

def run_custom_screen(
    preset_name: str,
) -> pd.DataFrame:

    """
    Run the financial screener using the configured thresholds.
    """

    config = load_screener_config()

    df = load_financial_ratios()

    df = calculate_composite_score(df)

    df = calculate_sector_relative_score(df)

    # Keep only the latest financial year for each company
    df = (
    df.sort_values("year")
      .groupby("company_id", as_index=False)
      .tail(1)
      .reset_index(drop=True)
)

    logger.info(
    "Latest-year dataset contains %d companies.",
    len(df),
)

    preset = config["presets"].get(preset_name)

    if preset is None:
        raise ValueError(
            f"Preset '{preset_name}' not found."
        )
    
    filters = preset.get("filters", {}).copy()

    de_filter = filters.pop("debt_to_equity", None)
    icr_filter = filters.pop("interest_coverage", None)

    filtered_df = apply_filters(
        df=df,
        filters=filters,
    )

    # Apply Debt-to-Equity rule
    if de_filter and "max" in de_filter:
        filtered_df = apply_de_filter(
            filtered_df,
            de_filter["max"],
        )

    # Apply Interest Coverage rule
    if icr_filter and "min" in icr_filter:
        filtered_df = apply_icr_filter(
            filtered_df,
            icr_filter["min"],
        )

    filtered_df = filtered_df.sort_values(
        by="composite_quality_score",
        ascending=False,
    ).reset_index(
        drop=True,
    )

    logger.info(
        "Screening completed. %d companies matched.",
        len(filtered_df),
    )

    return filtered_df

def export_screener_output(
    preset_results: dict,
) -> None:
    """
    Export all screener presets to a single Excel workbook.

    One worksheet is created for each preset.
    """

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for preset_name, df in preset_results.items():

            export_df = df.copy()

            available_columns = [
                column
                for column in EXPORT_COLUMNS
                if column in export_df.columns
            ]

            export_df = export_df[
                available_columns
            ]

            export_df.to_excel(
                writer,
                sheet_name=preset_name,
                index=False,
            )

            worksheet = writer.sheets[preset_name]

            apply_excel_formatting(
                writer=writer,
                preset_name=preset_name,
                worksheet=worksheet,
)

    logger.info(
        "Exported screener workbook: %s",
        OUTPUT_PATH,
    )

def apply_excel_formatting(writer, preset_name, worksheet):

    """Applies column widths, header styling, and conditional formatting to the exported Excel output."""
    config = load_screener_config()

    preset = config["presets"][preset_name]

    filters = preset.get("filters", {})

    headers = {}

    for cell in worksheet[1]:
        headers[cell.value] = cell.column

    for column_name, limits in filters.items():

        if column_name not in headers:
            continue

        excel_column = headers[column_name]

    for row in range(2, worksheet.max_row + 1):

            cell = worksheet.cell(
                row=row,
                column=excel_column,
            )

            value = cell.value

            if value is None:
                continue

            if "min" in limits:

                if value >= limits["min"]:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL

            elif "max" in limits:

                if value <= limits["max"]:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL
# ---------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------

if __name__ == "__main__":

    presets = [
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_blue_chip",
        "turnaround_watch",
    ]

    all_results = {}

    for preset in presets:
        print(f"\n{'='*60}")
        print(f"Preset : {preset}")
        print("="*60)

        result = run_custom_screen(preset)
        all_results[preset] = result

        print(f"Companies matched : {len(result)}")
        print(
    result[
        [
            "company_id",
            "broad_sector",
            "return_on_equity_pct",
            "debt_to_equity",
            "composite_quality_score",
        ]
    ].head(10)
)

    export_screener_output(all_results)
    print("\nExcel workbook generated successfully.")
        

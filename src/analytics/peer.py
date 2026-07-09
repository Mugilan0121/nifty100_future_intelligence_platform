"""
Peer Percentile Rankings.

Sprint 3 - Day 18

Loads peer groups and computes percentile rankings
within each peer group.
"""

import logging
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------
# Project Paths
# ---------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = PROJECT_ROOT / "nifty100.db"

PEER_GROUPS_PATH = (
    PROJECT_ROOT
    / "data"
    / "supporting"
    / "peer_groups.xlsx"
)

OUTPUT_PATH = (
    PROJECT_ROOT
    / "output"
    / "peer_comparison.xlsx"
)

# ---------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s - %(message)s"
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------
# Metrics to Rank
# ---------------------------------------------------------------------

RANKING_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]

# ---------------------------------------------------------------------
# Load Peer Groups
# ---------------------------------------------------------------------

def load_peer_groups():
    """
    Load peer group assignments from Excel.
    """

    peer_df = pd.read_excel(PEER_GROUPS_PATH)

    logger.info(
        "Loaded %d peer group assignments.",
        len(peer_df)
    )

    return peer_df

# ---------------------------------------------------------------------
# Load Financial Ratios
# ---------------------------------------------------------------------

def load_financial_ratios(conn):
    """
    Load financial ratios from SQLite.
    """

    query = """
    SELECT *
    FROM financial_ratios
    """

    ratios_df = pd.read_sql(query, conn)

    logger.info(
        "Loaded %d financial ratio records.",
        len(ratios_df)
    )

    return ratios_df

# ---------------------------------------------------------------------
# Load Company Names
# ---------------------------------------------------------------------

def load_company_names(conn):
    """
    Load company names from SQLite.
    """

    query = """
    SELECT
        id AS company_id,
        company_name
    FROM companies
    """

    company_df = pd.read_sql(
        query,
        conn,
    )

    logger.info(
        "Loaded %d companies.",
        len(company_df),
    )

    return company_df

# ---------------------------------------------------------------------
# Create SQLite Table
# ---------------------------------------------------------------------

def create_peer_percentiles_table(conn):
    """
    Create peer_percentiles table.
    """

    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS peer_percentiles (

            company_id TEXT,

            peer_group_name TEXT,

            metric TEXT,

            value REAL,

            percentile_rank REAL,

            year TEXT
        )
    """)

    conn.commit()

    logger.info("peer_percentiles table created successfully.")

# ---------------------------------------------------------------------
# Merge Peer Groups with Financial Ratios
# ---------------------------------------------------------------------

def calculate_peer_percentiles(peer_df, ratios_df):
    """
    Merge peer group assignments with financial ratios.
    """

    merged_df = ratios_df.merge(
        peer_df[
            [
                "company_id",
                "peer_group_name",
            ]
        ],
        on="company_id",
        how="left",
    )

    # Handle companies not assigned to a peer group
    merged_df["peer_group_name"] = merged_df[
        "peer_group_name"
    ].fillna("No peer group assigned")

    results = []

    valid_peer_df = merged_df[
        merged_df["peer_group_name"] !=
        "No peer group assigned"
    ].copy()

    for peer_group, group_df in valid_peer_df.groupby(
        "peer_group_name"
    ):
     
     for metric in RANKING_METRICS:

        percentile = group_df[metric].rank(
             pct=True,
             method="average",
         )

        if metric == "debt_to_equity":
             percentile = 1 - percentile

        for index, row in group_df.iterrows():

            results.append(
                {
                    "company_id": row["company_id"],
                    "peer_group_name": row["peer_group_name"],
                    "metric": metric,
                    "value": row[metric],
                    "percentile_rank": percentile.loc[index],
                    "year": row["year"],
                }
            )
    
    percentile_df = pd.DataFrame(results)

    logger.info(
        "Calculated %d percentile records.",
        len(percentile_df),
    )   

    return percentile_df

# ---------------------------------------------------------------------
# Save Percentiles
# ---------------------------------------------------------------------

def save_peer_percentiles(conn, percentile_df):
    """
    Save percentile rankings into SQLite.
    """

    cursor = conn.cursor()

    # Clear existing records
    cursor.execute("DELETE FROM peer_percentiles")

    percentile_df.to_sql(
        "peer_percentiles",
        conn,
        if_exists="append",
        index=False,
    )

    conn.commit()

    logger.info(
        "Saved %d percentile records to peer_percentiles table.",
        len(percentile_df),
    )

# ---------------------------------------------------------------------
# Build Peer Comparison Report
# ---------------------------------------------------------------------

def build_peer_comparison(
    peer_df,
    company_df,
    ratios_df,
    percentile_df,
):
    """
    Build one dataframe for each peer group.
    """

    latest_ratios = (
        ratios_df
        .sort_values(
            ["company_id", "year"]
        )
        .drop_duplicates(
            subset=["company_id"],
            keep="last",
        )
    )

    latest_ratios = latest_ratios.merge(
        peer_df[
            [
                "company_id",
                "peer_group_name",
            ]
        ],
        on="company_id",
        how="left",
    )

    latest_percentiles = (
        percentile_df
        .sort_values(
            ["company_id", "year"]
        )
        .drop_duplicates(
            subset=[
                "company_id",
                "metric",
            ],
            keep="last",
        )
    )

    pivot_percentiles = (
        latest_percentiles
        .pivot(
            index="company_id",
            columns="metric",
            values="percentile_rank",
        )
        .reset_index()
    )

    pivot_percentiles.columns = [

    (
        f"{col}_percentile"
        if col != "company_id"
        else col
    )

    for col in pivot_percentiles.columns
]

    peer_reports = {}

    for peer_group in peer_df[
        "peer_group_name"
    ].unique():

        members = peer_df[
            peer_df["peer_group_name"] == peer_group
        ]["company_id"]

        report_df = latest_ratios[
            latest_ratios["company_id"].isin(members)
        ].copy()

        report_df = report_df.merge(
            company_df,
            on="company_id",
            how="left",
        )

        report_df = report_df.merge(
            pivot_percentiles,
            on="company_id",
            how="left",
        )

        peer_reports[peer_group] = report_df

        logger.info(
            "%s : %d companies",
            peer_group,
            len(report_df),
        )

    return peer_reports

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966",
)

MEDIAN_FILL = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3",
)

def export_peer_comparison(peer_reports):

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for peer_group, report_df in peer_reports.items():

            sheet_name = peer_group[:31]

            report_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

            ws = writer.sheets[sheet_name]

            # -----------------------------
            # Benchmark company
            # -----------------------------

            if "composite_score" in report_df.columns:
                benchmark = report_df["composite_score"].idxmax()
            else:
                benchmark = report_df[
                    "return_on_equity_pct"
                ].idxmax()

            excel_row = benchmark + 2

            for cell in ws[excel_row]:
                cell.fill = BENCHMARK_FILL
                cell.font = Font(bold=True)

            # -----------------------------
            # Percentile colouring
            # -----------------------------

            percentile_cols = [
                i + 1
                for i, col in enumerate(report_df.columns)
                if col.endswith("_percentile")
            ]

            for col in percentile_cols:

                for row in range(2, len(report_df) + 2):

                    cell = ws.cell(
                        row=row,
                        column=col,
                    )

                    try:
                        value = float(cell.value)
                    except (TypeError, ValueError):
                        continue

                    if value >= 0.75:
                        cell.fill = GREEN_FILL

                    elif value <= 0.25:
                        cell.fill = RED_FILL

                    else:
                        cell.fill = YELLOW_FILL

            # -----------------------------
            # Median Row
            # -----------------------------

            median_row = ["Peer Median"]

            for col in report_df.columns[1:]:

                if pd.api.types.is_numeric_dtype(
                    report_df[col]
                ):
                    median_row.append(report_df[col].median())
                else:
                    median_row.append("")

            ws.append(median_row)

            last = ws.max_row

            for cell in ws[last]:
                cell.fill = MEDIAN_FILL
                cell.font = Font(bold=True)

            # -----------------------------
            # Autosize columns
            # -----------------------------

            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = length + 3

    logger.info(
        "Peer comparison exported to %s",
        OUTPUT_PATH,
    )

# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main():

    conn = sqlite3.connect(DATABASE_PATH)

    create_peer_percentiles_table(conn)

    peer_df = load_peer_groups()

    company_df = load_company_names(conn)

    ratios_df = load_financial_ratios(conn)

    percentile_df = calculate_peer_percentiles(
        peer_df,
        ratios_df,
    )

    save_peer_percentiles(
        conn,
        percentile_df,
    )

    peer_reports = build_peer_comparison(
        peer_df,
        company_df,
        ratios_df,
        percentile_df,
    )

    export_peer_comparison(
        peer_reports,
    )

    for name, df in peer_reports.items():

        print(f"\n{name}")

        print(df.head())

    print("\nPeer Percentiles")
    print(percentile_df.head(15))

    conn.close()


if __name__ == "__main__":
    main()